"""Excel report generation for NmapFusion - WITH PROPER SORTING"""

from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import ipaddress

class ExcelOutput:
    """Generate Excel compliance reports - WITH PROPER SORTING"""
    
    def __init__(self, output_dir):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.wb = Workbook()
        
        # Styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="00bcd4", end_color="00bcd4", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate(self, analysis_results, fusion_summary, nmap_commands, selected_tables):
        """Generate Excel report with NmapFusion branding - WITH PROPER SORTING"""
        
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Enforce correct table order
        table_order = ['table1', 'table2', 'table3', 'table4']
        
        for table in table_order:
            if table in selected_tables:
                if table == 'table1':
                    self._create_table1_sheet(analysis_results['table1'])
                elif table == 'table2':
                    self._create_table2_sheet(analysis_results['table2'])
                elif table == 'table3':
                    self._create_table3_sheet(analysis_results['table3'])
                elif table == 'table4':
                    self._create_table4_sheet(analysis_results['table4'])
        
        # Always generate these sheets
        self._create_nmap_command_sheet(nmap_commands)
        self._create_nse_findings_sheet(analysis_results['sorted_hosts'])
        self._create_subnets_sheet(analysis_results['sorted_hosts'])
        self._create_raw_data_sheet(analysis_results['sorted_hosts'])
        
        # Pass table3_data safely
        table3_data = analysis_results.get('table3', [])
        self._create_executive_summary_sheet(
            analysis_results['sorted_hosts'], 
            fusion_summary,
            table3_data
        )
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'nmapfusion_report_{timestamp}.xlsx'
        output_file = self.output_dir / filename
        
        # Save workbook
        self.wb.save(output_file)
        
        return output_file
    
    def _create_table1_sheet(self, table1_data):
        """Create Table 1: Host Summary Overview"""
        ws = self.wb.create_sheet("1_Host_Summary_Overview")
        
        headers = ["IP Address", "Hostname", "Total Ports", "TCP", "UDP", "Services", "OS"]
        ws.append(headers)
        
        # Format headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Add data
        if table1_data and isinstance(table1_data, list):
            for row in table1_data:
                ws.append([
                    row.get('ip', 'N/A'),
                    row.get('hostname', '-'),
                    row.get('total_ports', 0),
                    row.get('tcp_ports', 0),
                    row.get('udp_ports', 0),
                    row.get('total_services', 0),
                    row.get('os', 'unknown')[:30] if row.get('os') != 'unknown' else '-'
                ])
        
        self._adjust_column_widths(ws)
    
    def _create_table2_sheet(self, table2_data):
        """Create Table 2: Host Detailed Analysis"""
        ws = self.wb.create_sheet("2_Host_Detailed_Analysis")
        
        headers = ["IP Address", "Hostname", "OS", "Port", "Protocol", "Service", 
                  "Version", "NSE Findings", "Business Function"]
        ws.append(headers)
        
        # Format headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Add data
        if table2_data and isinstance(table2_data, dict):
            for ip, data in table2_data.items():
                if not data.get('ports'):
                    ws.append([
                        ip,
                        data.get('hostname', '-'),
                        data.get('os', 'unknown'),
                        '-', '-', '-', '-', '-', '-'
                    ])
                else:
                    for port in data.get('ports', []):
                        nse_summary = '; '.join(port.get('nse_summary', [])[:2]) if port.get('nse_summary') else '-'
                        ws.append([
                            ip,
                            data.get('hostname', '-'),
                            data.get('os', 'unknown'),
                            port.get('port', 'N/A'),
                            port.get('protocol', 'tcp'),
                            port.get('service', 'unknown'),
                            port.get('version', 'unknown')[:50] if port.get('version') != 'unknown' else '-',
                            nse_summary,
                            port.get('business_function', 'other').replace('_', ' ').title()
                        ])
        
        self._adjust_column_widths(ws)
    
    def _create_table3_sheet(self, table3_data):
        """Create Table 3: Port Frequency Distribution - SORTED BY PORT"""
        ws = self.wb.create_sheet("3_Port_Frequency_Distribution")
        
        headers = ["Port", "Protocol", "Host Count", "Service", "IPs"]
        ws.append(headers)
        
        # Format headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Add data if available
        if table3_data and isinstance(table3_data, list):
            for row in table3_data:
                ip_sample = ', '.join(row.get('ip_list', [])[:200]) if row.get('ip_list') else '—'
                if row.get('ip_list') and len(row['ip_list']) > 200:
                    ip_sample += f" +{len(row['ip_list'])-200} more"
                
                ws.append([
                    row.get('port', 'N/A'),
                    row.get('protocol', 'tcp'),
                    row.get('count', 0),
                    row.get('service', 'unknown'),
                    ip_sample
                ])
        
        self._adjust_column_widths(ws)
    
    def _create_table4_sheet(self, table4_data):
        """Create Table 4: Service Exposure Matrix - SORTED BY PORT - WITH PROPER IP GROUPING"""
        ws = self.wb.create_sheet("4_Service_Exposure_Matrix")
        
        headers = ["Port", "Protocol", "Exposed Hosts", "IP Addresses", 
                  "Hostnames", "OS", "Service Version", "Business Function"]
        ws.append(headers)
        
        # Format headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        # Add data if available
        if table4_data and isinstance(table4_data, dict):
            # Sort ports by port number
            sorted_ports = sorted(table4_data.items(), key=lambda x: x[1].get('port', 0))
            
            for port_key, data in sorted_ports:
                if data.get('hosts'):
                    # Collect all IPs, hostnames, etc. for this port
                    ip_list = []
                    hostname_list = []
                    os_list = []
                    version_list = []
                    function_list = []
                    
                    for host in data.get('hosts', []):
                        ip_list.append(host.get('ip', 'N/A'))
                        hostname = host.get('hostname', '-')
                        if hostname and hostname != '-':
                            hostname_list.append(hostname)
                        os = host.get('os', 'unknown')
                        if os != 'unknown':
                            os_list.append(os[:20])
                        version = host.get('version', '-')
                        if version and version != '-':
                            version_list.append(version[:50])
                        function = host.get('business_function', 'other').replace('_', ' ').title()
                        if function and function != 'Other':
                            function_list.append(function)
                    
                    # Format IP list with count
                    ip_display = ', '.join(ip_list[:20])
                    if len(ip_list) > 20:
                        ip_display += f" +{len(ip_list)-20} more"
                    
                    # Format hostnames (show unique)
                    unique_hostnames = list(dict.fromkeys(hostname_list))[:20]
                    hostname_display = ', '.join(unique_hostnames)
                    if len(unique_hostnames) > 20:
                        hostname_display += f" +{len(hostname_list)-20} more"
                    elif not hostname_display:
                        hostname_display = '-'
                    
                    # Format OS (show most common or unique)
                    unique_os = list(dict.fromkeys(os_list))[:2]
                    os_display = ', '.join(unique_os)
                    if len(unique_os) > 2:
                        os_display += f" +{len(os_list)-2} more"
                    elif not os_display:
                        os_display = '-'
                    
                    # Format versions (show unique)
                    unique_versions = list(dict.fromkeys(version_list))[:2]
                    version_display = ', '.join(unique_versions)
                    if len(unique_versions) > 2:
                        version_display += f" +{len(version_list)-2} more"
                    elif not version_display:
                        version_display = '-'
                    
                    # Format business functions (show unique)
                    unique_functions = list(dict.fromkeys(function_list))[:2]
                    function_display = ', '.join(unique_functions)
                    if len(unique_functions) > 2:
                        function_display += f" +{len(function_list)-2} more"
                    elif not function_display:
                        function_display = 'other'
                    
                    # Add single row for this port with all IPs
                    ws.append([
                        data.get('port', 'N/A'),
                        data.get('protocol', 'tcp'),
                        data.get('host_count', 0),
                        ip_display,
                        hostname_display,
                        os_display,
                        version_display,
                        function_display
                    ])
        
        self._adjust_column_widths(ws)
    
    def _create_nmap_command_sheet(self, nmap_commands):
        """Create Nmap Command sheet"""
        ws = self.wb.create_sheet("Scan_Configuration")
        
        ws['A1'] = "Nmap Command(s) Used"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = self.header_fill
        
        if nmap_commands and isinstance(nmap_commands, list):
            for i, cmd in enumerate(nmap_commands, 2):
                ws[f'A{i}'] = cmd
                ws[f'A{i}'].font = Font(name='Courier New')
        
        ws.column_dimensions['A'].width = 100
    
    def _create_nse_findings_sheet(self, hosts):
        """Create NSE Findings sheet"""
        ws = self.wb.create_sheet("NSE_Findings")
        
        headers = ["IP Address", "Hostname", "Script", "Output", "Port"]
        ws.append(headers)
        
        # Format headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Add NSE findings
        if hosts and isinstance(hosts, list):
            for host in hosts:
                for nse in host.get('nse', []):
                    ws.append([
                        host.get('ip', 'N/A'),
                        host.get('hostname', '-'),
                        nse.get('id', ''),
                        nse.get('output', '')[:200],
                        '-'
                    ])
        
        self._adjust_column_widths(ws)
    
    def _create_subnets_sheet(self, hosts):
        """Create Subnets sheet"""
        ws = self.wb.create_sheet("Subnet_Summary")
        
        from collections import defaultdict
        subnets = defaultdict(list)
        
        if hosts and isinstance(hosts, list):
            for host in hosts:
                subnet = host.get('subnet', 'unknown')
                subnets[subnet].append(host.get('ip', '0.0.0.0'))
        
        # Sort IPs within each subnet
        for subnet in subnets:
            try:
                subnets[subnet] = sorted(subnets[subnet], key=lambda x: ipaddress.ip_address(x))
            except:
                subnets[subnet] = sorted(subnets[subnet])
        
        headers = ["Subnet", "Host Count", "IP Range"]
        ws.append(headers)
        
        # Format headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Add subnet data
        for subnet, ips in sorted(subnets.items()):
            ws.append([
                subnet,
                len(ips),
                f"{ips[0]} - {ips[-1]}" if ips else '-'
            ])
        
        self._adjust_column_widths(ws)
    
    def _create_raw_data_sheet(self, hosts):
        """Create Raw Data sheet"""
        ws = self.wb.create_sheet("Raw_Data")
        
        headers = ["IP Address", "Hostname", "OS", "Port", "Protocol", "Service", 
                  "Version", "CVEs", "Source Files"]
        ws.append(headers)
        
        # Format headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Add raw data
        if hosts and isinstance(hosts, list):
            # Sort hosts by IP
            try:
                sorted_hosts = sorted(hosts, key=lambda x: ipaddress.ip_address(x.get('ip', '0.0.0.0')))
            except:
                sorted_hosts = hosts
            
            for host in sorted_hosts:
                cve_list = ', '.join([c.get('id', '') for c in host.get('cves', [])[:3]])
                source_files = ', '.join([Path(f).name for f in host.get('source_files', [])[:2]])
                
                if not host.get('ports'):
                    ws.append([
                        host.get('ip', 'N/A'),
                        host.get('hostname', '-'),
                        host.get('os', 'unknown'),
                        '-', '-', '-', '-',
                        cve_list,
                        source_files
                    ])
                else:
                    # Sort ports by port number
                    sorted_ports = sorted(host.get('ports', []), key=lambda x: x.get('port', 0))
                    for port in sorted_ports:
                        ws.append([
                            host.get('ip', 'N/A'),
                            host.get('hostname', '-'),
                            host.get('os', 'unknown'),
                            port.get('port', 'N/A'),
                            port.get('protocol', 'tcp'),
                            port.get('service', 'unknown'),
                            port.get('version', 'unknown')[:50],
                            cve_list,
                            source_files
                        ])
        
        self._adjust_column_widths(ws)
    
    def _create_executive_summary_sheet(self, hosts, fusion_summary, table3_data=None):
        """Create Executive Summary sheet with port statistics"""
        ws = self.wb.create_sheet("0_Executive_Summary")
        
        # Title
        ws['A1'] = 'NmapFusion - NETWORK ASSESSMENT SUMMARY'
        ws['A1'].font = Font(bold=True, size=16, color='00bcd4')
        ws.merge_cells('A1:F1')
        
        ws['A2'] = 'Enterprise Network Assessment Report'
        ws['A2'].font = Font(size=12, color='8899aa')
        ws.merge_cells('A2:F2')
        
        # Report Info
        ws['A4'] = 'Report Generated:'
        ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'].font = Font(bold=True)
        ws['B4'].font = Font(name='Courier New')
        
        ws['A5'] = 'Tool Version:'
        ws['B5'] = 'NmapFusion v1.0'
        ws['A5'].font = Font(bold=True)
        
        ws['A6'] = 'Files Analyzed:'
        ws['B6'] = fusion_summary.get('files_processed', 0)
        ws['A6'].font = Font(bold=True)
        
        # Statistics
        ws['A8'] = 'ASSESSMENT STATISTICS'
        ws['A8'].font = Font(bold=True, size=12, color='00bcd4')
        
        total_ports = 0
        total_cves = 0
        total_weak = 0
        
        if hosts and isinstance(hosts, list):
            total_ports = sum(len(h.get('ports', [])) for h in hosts)
            total_cves = sum(len(h.get('cves', [])) for h in hosts)
            total_weak = sum(len(h.get('weak_ciphers', [])) for h in hosts)
        
        # Safely get port range
        port_range = 'N/A'
        unique_ports = 0
        if table3_data and isinstance(table3_data, list) and len(table3_data) > 0:
            unique_ports = len(table3_data)
            try:
                first_port = table3_data[0].get('port', 'N/A')
                last_port = table3_data[-1].get('port', 'N/A')
                port_range = f"{first_port} - {last_port}"
            except (IndexError, KeyError, TypeError, AttributeError):
                port_range = 'N/A'
        
        stats = [
            ['Total Hosts Analyzed:', len(hosts) if hosts else 0],
            ['Total Open Ports:', total_ports],
            ['Total Unique Ports:', unique_ports],
            ['Port Range:', port_range],
            ['Total CVEs Found:', total_cves],
            ['Weak Cipher Suites:', total_weak],
            ['Total NSE Scripts:', fusion_summary.get('nse_merged', 0)]
        ]
        
        row_num = 10
        for i, stat in enumerate(stats):
            ws[f'A{row_num + i}'] = stat[0]
            ws[f'B{row_num + i}'] = stat[1]
            ws[f'A{row_num + i}'].font = Font(bold=True)
        
        self._adjust_column_widths(ws)
    
    def _adjust_column_widths(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell and cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 60)
            worksheet.column_dimensions[column_letter].width = adjusted_width
